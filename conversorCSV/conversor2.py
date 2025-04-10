import sys
import pandas as pd
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QLabel, QFileDialog, QMessageBox, QLineEdit
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

class CSVToExcelApp(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Conversor TXT/CSV para Excel')
        self.setGeometry(100, 100, 450, 400)
        self.setStyleSheet("""
            QWidget {
                background-color: #2C3E50;
                font-family: Arial;
            }
            QLabel {
                color: #ECF0F1;
                font-size: 14px;
                padding: 10px;
            }
            QLineEdit {
                padding: 5px;
                border: 1px solid #7F8C8D;
                border-radius: 5px;
                font-size: 14px;
            }
            QPushButton {
                background-color: #1ABC9C;
                color: white;
                font-size: 16px;
                padding: 10px;
                border: none;
                border-radius: 10px;
            }
            QPushButton:hover {
                background-color: #16A085;
            }
            QPushButton:disabled {
                background-color: #7F8C8D;
                color: #BDC3C7;
            }
        """)

        layout = QVBoxLayout()

        self.instructions_label = QLabel('Carregue um arquivo TXT/CSV para converter em Excel', self)
        self.instructions_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.instructions_label)

        self.delimiter_label = QLabel('Delimitador (padrão: ","):', self)
        layout.addWidget(self.delimiter_label)

        self.delimiter_input = QLineEdit(self)
        self.delimiter_input.setPlaceholderText('Informe o delimitador, ex: "," ou ";"')
        layout.addWidget(self.delimiter_input)

        self.upload_button = QPushButton('Upload TXT/CSV', self)
        self.upload_button.clicked.connect(self.process_csv)
        layout.addWidget(self.upload_button)

        self.status_label = QLabel('', self)
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_label)

        self.save_button = QPushButton('Salvar como Excel', self)
        self.save_button.setEnabled(False)
        self.save_button.clicked.connect(self.save_excel)
        layout.addWidget(self.save_button)

        self.setLayout(layout)

    def process_csv(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Selecione o arquivo TXT/CSV", "", "Arquivos de Texto (*.txt *.csv)")
        if not file_path:
            return

        self.status_label.setText(f"Arquivo carregado: {file_path}")
        self.save_button.setEnabled(True)
        self.csv_path = file_path

    def save_excel(self):
        try:
            delimiter = self.delimiter_input.text() or ','
            df = pd.read_csv(self.csv_path, delimiter=delimiter)

            excel_path, _ = QFileDialog.getSaveFileName(self, "Salvar como Excel", "", "Excel Files (*.xlsx)")
            if not excel_path:
                return

            # Salva o DataFrame no Excel
            df.to_excel(excel_path, index=False)

            # Abre o arquivo Excel para adicionar uma linha em branco no topo
            wb = load_workbook(excel_path)
            ws = wb.active

            # Adiciona uma linha em branco no topo (linha 1)
            ws.insert_rows(1)

            # Estiliza a primeira linha (cabeçalho) para não ser em negrito, sem bordas e alinhado à esquerda
            for cell in ws[2]:
                # Remover negrito
                cell.font = Font(bold=False)
                # Remover bordas
                cell.border = Border(left=Side(border_style=None),
                                    right=Side(border_style=None),
                                    top=Side(border_style=None),
                                    bottom=Side(border_style=None))
                # Alinhar à esquerda
                cell.alignment = Alignment(horizontal="left")

            # Salva as alterações no arquivo
            wb.save(excel_path)

            self.status_label.setText(f"Arquivo salvo como: {excel_path}")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar o arquivo: {str(e)}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = CSVToExcelApp()
    window.show()
    sys.exit(app.exec_())
